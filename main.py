import wx
import wx.grid
from datetime import datetime
from mygrid import MyGrid
from openpyxl import load_workbook, Workbook
import os
from dates import getInvalidDates, get_some_years, turnAroundDate
from bd import makeQueryMandanteCalendario, checkAllDates



class MainFrame(wx.Frame):
    """
    A Frame that says Hello World
    """

    def __init__(self, *args, **kw):
        # ensure the parent's __init__ is called
        super(MainFrame, self).__init__(*args, **kw)
        
        self.makeTabs()

        # and a status bar
        # self.CreateStatusBar()
        # self.SetStatusText("Working!")

    def makeVboxLines(self, panel):
        self.input_text_area = wx.TextCtrl(panel, style=wx.TE_MULTILINE, size=(300, 200))
        self.output_text_area = wx.TextCtrl(panel, style=wx.TE_MULTILINE, size=(300, 200))
        
        self.btn_all = wx.Button(panel, label="> Todo >")
        self.btn_all.Bind(wx.EVT_BUTTON, lambda event: self.textTransformations(event, 0))

        self.btn_commas = wx.Button(panel, label="> Pasar a comas >")
        self.btn_commas.Bind(wx.EVT_BUTTON, lambda event: self.textTransformations(event, 1))

        self.btn_unique = wx.Button(panel, label="> Unico >")
        self.btn_unique.Bind(wx.EVT_BUTTON, lambda event: self.textTransformations(event, 2))

        max_width = max(self.btn_all.GetBestSize().width, self.btn_commas.GetBestSize().width, self.btn_unique.GetBestSize().width)
        self.btn_all.SetMinSize(wx.Size(max_width, self.btn_all.GetBestSize().height))
        self.btn_commas.SetMinSize(wx.Size(max_width, self.btn_commas.GetBestSize().height))
        self.btn_unique.SetMinSize(wx.Size(max_width, self.btn_unique.GetBestSize().height))

        vbox_buttons = wx.BoxSizer(wx.VERTICAL)
        vbox_buttons.Add(self.btn_all, flag=wx.ALL | wx.CENTER, border=10)
        vbox_buttons.Add(self.btn_commas, flag=wx.ALL | wx.CENTER, border=10)
        vbox_buttons.Add(self.btn_unique, flag=wx.ALL | wx.CENTER, border=10)
        
        vbox = wx.BoxSizer(wx.HORIZONTAL)
        vbox.Add(self.input_text_area, proportion = 1, flag = wx.EXPAND | wx.ALL, border = 10 )
        vbox.Add(vbox_buttons, flag=wx.ALL | wx.EXPAND, border=10 )
        vbox.Add(self.output_text_area, proportion = 1, flag = wx.EXPAND | wx.ALL, border = 10 )

        panel.SetSizer(vbox)

    def textTransformations(self, event, flag:int):
        # Retrieve the text from the text area
        input_text = self.input_text_area.GetValue()
        output_text = ""
        match flag:
            case 0:
                output_text = self.uniques(self.linesToCommas(input_text), ',')
            case 1:
                output_text = self.linesToCommas(input_text)
            case 2:
                output_text = self.uniques(input_text, '\n')
        self.output_text_area.SetValue(output_text)
        self.clipboardOutput(output_text)
        wx.MessageBox(f"Se ha copiado el resultado a tu portapapeles", "Datos transformados", wx.OK | wx.ICON_INFORMATION)


    def linesToCommas(self, text:str):
        return text.replace("\n", ",")
    
    def uniques(self, text:str, separation:str):
        unique_values = set(text.split(separation))
        output = ""
        for unique in unique_values:
            output += unique + separation
        return output.strip(separation)
    
    def makeVboxCalendar(self, panel):
        # Create a grid (spreadsheet) control
        #self.grid = gridlib.Grid(panel)
        self.grid = MyGrid(panel)
        self.grid.CreateGrid(24, 6)
        
        # Set the number of rows and columns
        #self.grid.CreateGrid(24, 6)  # 5 rows, 3 columns
        
        # Set column labels
        self.grid.SetColLabelValue(0, "Razon Social")
        self.grid.SetColLabelValue(1, "Fecha Pago")
        self.grid.SetColLabelValue(2, "Fecha Apertura")
        self.grid.SetColLabelValue(3, "Fecha Informe")
        self.grid.SetColLabelValue(4, "Fecha Cierre")
        self.grid.SetColLabelValue(5, "Mandante ID")
        self.grid.AutoSizeColumns()

        #self.grid.Bind(MyGrid.EVT_GRID_CELL_CHANGING, self.OnCellValueChanged)
        
        # Set some cell values
        #self.grid.SetCellValue(0, 0, "Alice")

        self.add_row_button = wx.Button(panel, label="Agregar Fila")
        self.add_row_button.Bind(wx.EVT_BUTTON, self.OnAddRow)
        
        # Add a button to add a column
        # self.add_col_button = wx.Button(panel, label="Agregar Columna")
        # self.add_col_button.Bind(wx.EVT_BUTTON, self.OnAddColumn)

        
        # Create a button to show the current grid data
        self.show_button = wx.Button(panel, label="Obtener SQL")
        self.show_button.Bind(wx.EVT_BUTTON, self.OnShowData)

        # Add a button to set a value to the selected range
        self.set_value_button = wx.Button(panel, label="Dar valor a selección")
        self.set_value_button.Bind(wx.EVT_BUTTON, self.OnSetValue)

        # Add a button to set a value to the selected range
        self.turnaround_date_button = wx.Button(panel, label="Dar vuelta fechas")
        self.turnaround_date_button.Bind(wx.EVT_BUTTON, self.OnTurnaroundDate)

        # Create a wx.Choice control
        self.listed_years = get_some_years()
        self.choice = wx.Choice(panel, choices=self.listed_years)
        self.choice.Bind(wx.EVT_CHOICE, self.OnChoice)
        self.choice.SetSelection(3)
        self.CreateStatusBar()
        self.SetStatusText(f"El año {self.listed_years[3]} se encuentra cargado para validar feriados")
        self.invalid_dates = getInvalidDates(self.listed_years[3])
        self.grid.invalid_dates = self.invalid_dates

        # Determining the max widht and height among buttons
        max_width = max(self.add_row_button.GetBestSize().width, self.set_value_button.GetBestSize().width, self.turnaround_date_button.GetBestSize().width, self.show_button.GetBestSize().width)
        # max_height = max(self.add_row_button.GetBestSize().height, self.set_value_button.GetBestSize().height, self.turnaround_date_button.GetBestSize().height, self.show_button.GetBestSize().width)

        self.add_row_button.SetMinSize(wx.Size(max_width, self.add_row_button.GetBestSize().height))
        self.set_value_button.SetMinSize(wx.Size(max_width, self.set_value_button.GetBestSize().height))
        self.turnaround_date_button.SetMinSize(wx.Size(max_width, self.turnaround_date_button.GetBestSize().height))
        self.show_button.SetMinSize(wx.Size(max_width, self.show_button.GetBestSize().height))
        self.choice.SetMinSize(wx.Size(max_width, self.choice.GetBestSize().height))


        vbox_buttons = wx.BoxSizer(wx.VERTICAL)
        vbox_buttons.Add(self.add_row_button, flag=wx.ALL | wx.CENTER, border=10)
        # vbox_buttons.Add(self.add_col_button, flag=wx.ALL | wx.CENTER, border=10)
        vbox_buttons.Add(self.set_value_button, flag=wx.ALL | wx.CENTER, border=5)
        vbox_buttons.Add(self.turnaround_date_button, flag=wx.ALL | wx.CENTER, border=5)
        vbox_buttons.Add(self.show_button, flag=wx.ALL | wx.CENTER, border=10)
        vbox_buttons.Add(self.choice, flag=wx.ALL | wx.CENTER, border=10)

        # Arrange controls in a vertical box sizer
        vbox = wx.BoxSizer(wx.HORIZONTAL)
        vbox.Add(self.grid, proportion=1, flag=wx.EXPAND | wx.ALL, border=10)
        vbox.Add(vbox_buttons, flag=wx.ALL | wx.EXPAND, border=10 )
        
        panel.SetSizer(vbox)

    def OnAddRow(self, event):
        # Get the current number of rows
        num_rows = self.grid.GetNumberRows()
        
        # Append a new row
        self.grid.AppendRows(1)
        
        # Optionally, you can initialize the new row with some default values
        # for col in range(self.grid.GetNumberCols()):
        #     self.grid.SetCellValue(num_rows, col, f"Row {num_rows} Col {col}")

    def OnAddColumn(self, event):
        # Get the current number of columns
        num_cols = self.grid.GetNumberCols()
        
        # Append a new column
        self.grid.AppendCols(1)
        
        # Set a label for the new column
        self.grid.SetColLabelValue(num_cols, f"Column {num_cols}")
        
        # Optionally, you can initialize the new column with some default values
        for row in range(self.grid.GetNumberRows()):
            self.grid.SetCellValue(row, num_cols, f"Row {row} Col {num_cols}")

    def OnSetValue(self, event):
        # Prompt the user for the value to set
        value_to_set = wx.GetTextFromUser("Enter the value to set:", "Set Value")
        if not value_to_set:
            return  # If the user cancels or enters an empty value, do nothing

        # Get the selected range
        selection = self.grid.get_selection()
        if not selection:
            return  # If no selection, do nothing

        start_row, start_col, end_row, end_col = selection

        # Set the value for each cell in the selected range
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                self.grid.SetCellValue(row, col, value_to_set)

        #wx.MessageBox("Value set successfully!", "Info", wx.OK | wx.ICON_INFORMATION)

    def OnTurnaroundDate(self, event):
        # Get the selected range
        selection = self.grid.get_selection()
        if not selection:
            return  # If no selection, do nothing

        start_row, start_col, end_row, end_col = selection

        # Set the value for each cell in the selected range
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                new_value = turnAroundDate(self.grid.GetCellValue(row, col))
                self.grid.SetCellValue(row, col, new_value)
                if new_value in self.invalid_dates:
                    bg_color = wx.Colour(255, 0, 0)  # Red background
                else:
                    bg_color = wx.Colour(255, 255, 255)  # White background
                # Create a custom renderer and set the background color
                self.grid.UpdateCellAttr(row, col, bg_color)


    def OnChoice(self, event):
        selection = self.choice.GetStringSelection()
        self.invalid_dates = getInvalidDates(selection)
        self.grid.invalid_dates = self.invalid_dates
        self.SetStatusText(f"El año {selection} se encuentra cargado para validar feriados")
        #wx.MessageBox(f"You selected: {selection}", "Info", wx.OK | wx.ICON_INFORMATION)


    def OnShowData(self, event):
        # Get data from the grid
        data = []
        saved = False
        for row in range(self.grid.GetNumberRows()):
            row_data = [self.grid.GetCellValue(row, col) for col in range(self.grid.GetNumberCols())]
            data.append(row_data)
        errors = checkAllDates(data)
        if len(errors) > 0:
            wx.MessageBox(f"Errores de logica de fechas en las filas :\n{errors}", "Errores de logica de fechas", wx.OK | wx.ICON_INFORMATION)  
        else:
            saved = makeQueryMandanteCalendario(data)
        if (saved):
            wx.MessageBox(f"Se ha guardado un archivo sql", "Data", wx.OK | wx.ICON_INFORMATION)
        # Display the data in a message box
        # data_str = "\n".join([", ".join(row) for row in data])
        # wx.MessageBox(f"Grid Data:\n{data_str}", "Data", wx.OK | wx.ICON_INFORMATION)

    def on_change(self, event):
        row = event.GetRow()
        col = event.GetCol()
        value = self.grid.GetCellValue(row, col)
        wx.MessageBox(f"You selected: {self.invalid_dates}", "Info", wx.OK | wx.ICON_INFORMATION)
        
        # Example condition: change background color if value contains '5'
        if value in self.invalid_dates:
            bg_color = wx.Colour(255, 0, 0)  # Red background
        else:
            bg_color = wx.Colour(255, 255, 255)  # White background
        # Create a custom renderer and set the background color
        self.grid.UpdateCellAttr(row, col, bg_color)
        # attr = MyGrid.GridCellAttr()
        # attr.SetBackgroundColour(bg_color)
        # self.grid.SetCellAttr(row, col, attr)

        #event.Skip()  # Continue to propagate the event

    def makeVboxExcel(self, panel):
        self.btn_load_excel = wx.Button(panel, label='Load Excel File')
        vbox_buttons = wx.BoxSizer(wx.VERTICAL)
        vbox_buttons.Add(self.btn_load_excel, flag=wx.ALL | wx.CENTER, border=10)

        self.btn_load_excel.Bind(wx.EVT_BUTTON, self.on_load_excel)

        self.text_ctrl = wx.TextCtrl(panel, style=wx.TE_MULTILINE | wx.TE_READONLY, size=(400, 300))
        vbox_buttons.Add(self.text_ctrl, flag=wx.ALL | wx.EXPAND, border=10)

        panel.SetSizer(vbox_buttons)

        


    def on_load_excel(self, event):
        wildcard = "Excel files (*.xlsx)|*.xlsx|All files (*.*)|*.*"
        dialog = wx.FileDialog(self, "Open Excel file", wildcard=wildcard, style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST)

        if dialog.ShowModal() == wx.ID_CANCEL:
            return  # User cancelled the file dialog

        # Get the selected file path
        path = dialog.GetPath()

        # Load the Excel file using openpyxl
        try:
            workbook = load_workbook(filename=path)
            sheet_name = 'Trabajadores'
            # sheet = workbook.active
            sheet = workbook[sheet_name]
            self.makeExcel(sheet, dialog.GetFilename(), sheet_name)
            # Read and display the content in the wx.TextCtrl
            # content = ""
            # for row in sheet.iter_rows(values_only=True):
            #     content += "\t".join([str(cell) if cell is not None else "" for cell in row]) + "\n"

            # self.text_ctrl.SetValue(content)
        except Exception as e:
            wx.LogError(f"Cannot open file '{os.path.basename(path)}'. Error: {str(e)}")

    def makeExcel(self, sheet, file_name, sheet_name):
        print("pre column headers")
        column_headers = {index: cell.value for index, cell in enumerate(sheet[1], start=1)}
        print("post column headers")
        expected_headers = {
            1 : 'Ítem',
            2 : 'País de origen',
            3 : 'Documento',
            4 : 'Número CI',
            5 : 'Nombre ',
            6 : 'Ap. Paterno ',
            7 : 'Ap. Materno',
            8 : 'F. Nacimiento',
            9 : 'Sexo',
            10 : 'Región',
            11 : 'Dirección',
            12 : 'Comuna',
            13 : 'Teléfono',
            14 : 'ISAPRE',
            15 : 'AFP',
            16 : 'Pensionado', 
            17 : 'Trabajador Dueño',
            18 : 'Email',
            19 : 'Discapacidad',
            20 : 'Cargo',
            21 : 'Artículo 22',
            22 : 'Contrato',
            23 : 'Fecha Inicio Contrato',
            24 : 'Fecha Término Contrato',
            25 : 'Jornada',
            26 : 'Fecha Ingreso a la Obra'
        }
        nullable = [
            1,
            24
        ]
        content = []
        errors = []
        for row_index, row in enumerate(sheet.iter_rows(min_row = 2, values_only=True), start = 2):
            processed_row = {}
            for index,cell in enumerate(row, start = 1):
                print("pre- if nullable")
                if (cell is None or cell == "") and index not in nullable: # Otras restricciones se pueden agregar aqui
                    print("in- if nullable")
                    errors.append(f"La columna {index} de la fila {row_index} por tener un valor nulo.")
                    continue
                else:
                    print("else- if nullable")
                    if isinstance(cell, datetime):
                        processed_row[column_headers[index]] = cell.date()
                    else:
                        processed_row[column_headers[index]] = cell
                print("post- if nullable")
            content.append(processed_row)
        if len(errors) > 0:
            print("pre-append print errors")
            self.text_ctrl.SetValue(self.printable_error(errors)) # Formatear errores para impresion
            self.write_excel(column_headers, content, "Con_errores_"+file_name, sheet_name)
            print("post-append print errors")
        else:
            print("pre- write excel")
            self.write_excel(column_headers, content, file_name, sheet_name)
            print("post- write excel")
            # processed_row = [cell if cell is not None and index not in (nullable) else "" for index, cell in enumerate(row)]
            # content += "\t".join(processed_row) + "\n"
    def printable_error (self, errors):
        printable = ""
        for error in errors:
            printable+= error+'\n'
        return printable
        
    def write_excel(self, column_headers, content, file_name, sheet_name):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name


        # Write column headers
        headers = [header for _, header in sorted(column_headers.items())]
        print("pre-append headers")
        sheet.append(headers)
        print("post-append headers")
        # Write data rows
        print(content)
        for row_data in content:
            row = [row_data.get(header, '') for _, header in sorted(column_headers.items())]
            print(row)
            print("pre-append row")
            sheet.append(row)
            print("post-append row")

        # Save the workbook
        workbook.save("cleaned_"+ file_name)
        self.text_ctrl.SetValue("Se ha guardado una versión limpiada en excel") # Formatear errores para impresion
        wx.MessageBox(f"Se ha guardado una versión limpiada en excel", "Archivo guardado", wx.OK | wx.ICON_INFORMATION)
        pass

    def clipboardOutput(self, text):
        if wx.TheClipboard.Open():
            wx.TheClipboard.Clear()
            wx.TheClipboard.SetData(wx.TextDataObject(text))
            wx.TheClipboard.Close()

    def makeTabs(self):
        notebook = wx.Notebook(self)

        tab1 = wx.Panel(notebook)
        notebook.AddPage(tab1, "Listados")

        self.makeVboxLines(tab1)

        tab2 = wx.Panel(notebook)
        notebook.AddPage(tab2, "Calendario de Cierre")

        self.makeVboxCalendar(tab2)

        tab3 = wx.Panel(notebook)
        notebook.AddPage(tab3, "Limpiar Excel")

        self.makeVboxExcel(tab3)

        # vbox2 = wx.BoxSizer(wx.VERTICAL)
        # label2 = wx.StaticText(tab2, label="This is the second tab")
        # vbox2.Add(label2, flag=wx.ALL, border = 10)
        # tab2.SetSizer(vbox2)

        sizer = wx.BoxSizer(wx.VERTICAL)
        sizer.Add(notebook, 1, wx.EXPAND)
        self.SetSizerAndFit(sizer)

        self.SetSize((400, 300))
        self.Centre()



if __name__ == '__main__':
    # When this module is run (not imported) then create the app, the
    # frame, show it, and start the event loop.
    app = wx.App()
    frm = MainFrame(None, title='SubcontrataLey Misc. Functions')
    frm.Show()
    app.MainLoop()